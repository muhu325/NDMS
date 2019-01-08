import os, json, time, re
from app01 import models
from app01.utils.others import my_read_very_eager, mydecode, port_mapping_reverse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from app01.utils.spider import get_busenesses_brief_info_by_businesses_code
from concurrent.futures import ThreadPoolExecutor
from app01.utils import network_device
from app01.utils.NDMS_settings import INT1, INT2


class Row:
    """每一个对象对应一行查询结果"""

    def __init__(self, num, row_excel):
        # self.keywords = []  # 用于查询的关键字,改主意了，把查询关键字放在onu对象中好了
        self.row_excel = row_excel  # excel的完整行内容
        self.num = num
        self.query_obj_list = []

    def get_query_objs(self):
        """根据excel表行内容进行匹配，将查询到的olt对象、onu对象、switch对象存放进self.query_obj_list中"""
        # 参数为割接表中每一行的信息
        # 返回值是一个列表，包含的元素为字典，里面包含了用于登录设备查询的信息
        # 例如{"query_type": "onu","device_type": version,"device_ip": ip,"onu_interface": onu_interface,"olt_interface": olt_interfce}

        regex1 = re.compile("\d{4}[A-Za-z]{1,2}-[A-Za-z]{1,2}")  # 匹配分光器
        regex2 = re.compile("\d{4}[A-Za-z]{1,2}-[A-Za-z]{1,2}-[A-Za-z]{1,2}")  # #匹配分光器，优先

        regex3 = re.compile(
            "\d{4}\S*?\.[^\t\n\r\f]{4,}?\.\S*")  # 可以匹配到1127-D-P-A.H3C12518.G17-0-15、1100-POP-01.S12508.GE7/0/42

        regex4 = re.compile("[A-Z]_[A-Z]\d{2}_\d{8}")
        regex5 = re.compile("Z_P_\d*")
        regex6 = re.compile("ZL_\d*")
        regex7 = re.compile("HZ_YWSL_{1,2}\d+")
        # regex8 = re.compile("\d{4}[IV]-[A-Za-z]-\S*?\d{1,2}[-/]\d{1,2}[-/]\d{1,2}")

        regex11 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([ONUonu]{3}[ \t]*?\d{1,2}/0/\d{1,2}:\d{1,2})")  # 华三onu
        regex12 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?[ONUonu]{3}[ \t]*?([123456789]/[123456789]/\d{1,2})")  # 瑞斯onu
        regex13 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([OLTolt]{3}[ \t]*?\d{1,2}/0/\d{1,2})")  # 华三olt
        regex14 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?[OLTolt]{3}[ \t]*?([123456789]/[123456789])")  # 瑞斯olt
        # ------新增了交换机的割接确认-----------
        regex15 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([ethETH]{3}[ \t]*?\d{1,2}/\d{1,2})")  # Cisco Nexus端口
        regex16 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([GgTtXExe]{2}-\d{1,2}/\d{1,2}/\d{1,2})")  # juniper xe/ge/et端口
        regex17 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Aaxe]{2}\d{1,2}\.{0,1}\d{0,4})")  # juniper ae端口
        regex18 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Bb][Uu][Nn][dD][Ll][Ee]\-[eE][therTHER]{0,4} {0,3}\d{1,2}\.{0,1}\d{0,4})")  # PE_9K bundle-ether端口
        regex19 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Ff][ \t]*?\d{1,2}/\d{1,2})")  # 思科交换机f口
        regex20 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([tT][enigENIG\-]{0,6}[ \t]{0,5}\d{1,2}/\d{1,2}/\d{1,2}/\d{1,2})")  # PE_9K Te端口
        regex21 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([tT][enigENIG\-]{0,6}[ \t]{0,5}\d{1,2}/\d{1,2}/\d{1,2})")  # 华三交换机ten口
        regex22 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Gg][igabherntIGABHERNT]{0,14}[ \t]{0,5}?\d{1,2}/\d{1,2}/\d{1,2}/\d{0,2})")  # 华三交换机g口,PE_9K Gi端口
        regex23 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Gg][igabherntIGABHERNT]{0,14}[ \t]{0,5}?\d{1,2}/\d{1,2}/\d{1,2})")  #
        regex24 = re.compile(
            r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[ \t]*?([Gg][igabherntIGABHERNT]{0,14}[ \t]{0,5}?\d{1,2}/\d{1,2})")  # 思科7609 Gi端口



        regex30 = re.compile(r"(\d{4}[IV][-_]\S)[^.]{0,10}\.[\w ]{0,10}\.([\w /]*)")  # 匹配交换机的配置信息（匹配excel文本）
        regex31 = re.compile(r"(\d{4}[IV][-_]\S)[^.]*\.[^.]*\.(\S*)")  # 匹配交换机的配置信息(匹配由爬虫爬出的配置信息）

        # 匹配优先级1：ip+端口，例如 10.0.0.1 olt2/0/1---regex11/12/13/14
        # 匹配优先级2：配置信息，例如 0227E-B.7602S.15---regex3
        # 匹配优先级3：业务编码，例如 D_D00_21025829---regex4/5/6/7
        # 匹配优先级4：分光器编号，例如 1127e-ar-------regex2/1
        # 高优先级匹配若有结果，则不继续向下匹配
        # 其实最根本的并不是匹配到就跳出并返回，而是匹配到后，继续查找有结果，才返回
        # 想到了，在每次循环前，判断一下list_re是否有值；若有，则跳过；若无，则进去匹配
        list_regex = [regex4, regex5, regex6, regex7]
        list_regex2 = [regex2, regex1]
        list_regex3 = [regex11, regex12, regex13, regex14]

        #匹配端口查询时，主要考虑到2种：比如mx960的xe-0/0/0这种形式的肯定是唯一的，所以匹配了，还继续匹配其他正则
        #但是 g0/0/0/0这种类型的端口，如果也继续匹配其他正则的话，一个g0/0/0/0还可以匹配出g0/0/0,g0/0，所以必须匹配即退出
        list_regex4 = [regex15, regex16, regex17, regex18, regex19,]
        list_regex5 = [regex20, regex21, ]
        list_regex6 = [regex22, regex23, regex24,]

        # regex_result = ""
        # 匹配华三onu:ip+onu2/0/1:1
        if not self.query_obj_list:
            regex_results = re.findall(regex11, self.row_excel)
            if regex_results:
                regex_results = list(set(regex_results))  # 去重
            for ip, onu_interface in regex_results:
                ip = ip.strip()
                onu_interface = onu_interface.lower().strip().replace(" ", "").replace("\t", "")
                olt_interface = "olt" + re.findall("\d{1,2}/0/\d{1,2}", onu_interface)[0]
                objs_onu = models.Onu.objects.filter(interface=onu_interface, olt__interface=olt_interface,
                                                     olt__epon__IP=ip)

                for obj_onu in objs_onu:
                    obj_onu.query_type = "onu"
                    obj_onu.version = obj_onu.olt.epon.tag.version.name
                    obj_onu.ip = ip
                    obj_onu.interfaces = onu_interface
                    obj_onu.keywords = "%s  %s" % (ip, onu_interface)
                    obj_onu.num = self.num

                    self.query_obj_list.append(obj_onu)
        # 匹配瑞斯onu:ip+onu 1/1/1
        if not self.query_obj_list:
            regex_results = re.findall(regex12, self.row_excel)
            if regex_results:
                regex_results = list(set(regex_results))  # 去重
            for ip, onu_interface in regex_results:
                ip = ip.strip()
                onu_interface = onu_interface.lower().strip().replace(" ", "").replace("\t", "")
                olt_interface = re.findall("([123456789]/[123456789])/", onu_interface)[0]

                objs_onu = models.Onu.objects.filter(interface=onu_interface, olt__interface=olt_interface,
                                                     olt__epon__IP=ip)
                for obj_onu in objs_onu:
                    obj_onu.query_type = "onu"
                    obj_onu.version = obj_onu.olt.epon.tag.version.name
                    obj_onu.ip = ip
                    obj_onu.interfaces = onu_interface
                    obj_onu.keywords = "%s  onu %s" % (ip, onu_interface)
                    obj_onu.num = self.num
                    self.query_obj_list.append(obj_onu)
        # 匹配华三olt:ip+olt 2/0/1
        if not self.query_obj_list:
            regex_results = re.findall(regex13, self.row_excel)
            if regex_results:
                # print("匹配华三olt：", ret1)
                regex_results = list(set(regex_results))  # 去重
            for ip, olt_interface in regex_results:
                ip = ip.strip()
                olt_interface = olt_interface.lower().strip().replace(" ", "").replace("\t", "")
                objs_olt = models.Olt.objects.filter(epon__IP=ip, )
                objs_olt = models.Olt.objects.filter(epon__IP=ip, interface=olt_interface)
                for obj_olt in objs_olt:
                    obj_olt.query_type = "olt"
                    obj_olt.version = obj_olt.epon.tag.version.name
                    obj_olt.ip = ip
                    obj_olt.interfaces = olt_interface  # 将要查询的端口插入到对象中
                    obj_olt.keywords = "%s  %s" % (ip, olt_interface)
                    obj_olt.num = self.num

                    self.query_obj_list.append(obj_olt)
        # 匹配瑞斯olt:ip+olt 1/1
        if not self.query_obj_list:
            regex_results = re.findall(regex14, self.row_excel)
            # print(regex_results)
            if regex_results:
                # print("匹配瑞斯olt：", ret1)
                regex_results = list(set(regex_results))  # 去重
            for ip, olt_interface in regex_results:
                ip = ip.strip()
                olt_interface = olt_interface.lower().strip().replace(" ", "").replace("\t", "")
                objs_olt = models.Olt.objects.filter(epon__IP=ip, interface=olt_interface)
                for obj_olt in objs_olt:
                    obj_olt.query_type = "olt"
                    obj_olt.version = obj_olt.epon.tag.version.name
                    obj_olt.ip = ip
                    obj_olt.interfaces = olt_interface  # 将要查询的端口插入到对象中
                    obj_olt.keywords = "%s  olt %s" % (ip, olt_interface)
                    obj_olt.num = self.num
                    self.query_obj_list.append(obj_olt)

        # -----新增了交换机的割接确认-----
        """这些待修改，因为交换机的modele还没写。。。"""
        # 用于匹配各种非OLT/ONU口
        if not self.query_obj_list:
            all_regex_results = []
            for regex in list_regex4:
                regex_results = re.findall(regex, self.row_excel)
                if regex_results:
                    for i in regex_results:
                        all_regex_results.append(i)
            for regex in list_regex5:
                regex_results = re.findall(regex, self.row_excel)
                if regex_results:
                    for i in regex_results:
                        all_regex_results.append(i)
                    break
            for regex in list_regex6:
                regex_results = re.findall(regex, self.row_excel)
                if regex_results:
                    for i in regex_results:
                        all_regex_results.append(i)
                    break
            if all_regex_results:
                all_regex_results = list(set(all_regex_results))  # 去重
            for ip, interface in all_regex_results:
                ip = ip.strip()
                interface = interface.lower().strip().replace(" ", "").replace("\t", "")
                objs_device = models.Device.objects.filter(IP=ip)
                for obj_device in objs_device:
                    obj_device.query_type = obj_device.tag.name  # 只有非EPON设备的查询类型用tag.name（POP/CMTS_POP/...)来命名
                    obj_device.version = obj_device.tag.version.name
                    obj_device.ip = ip
                    obj_device.interfaces = interface  # 将要查询的端口插入到对象中
                    obj_device.keywords = "%s  %s" % (ip, interface)
                    obj_device.num = self.num

                    self.query_obj_list.append(obj_device)
        # 匹配交换机配置信息 # 不好用啊。。。
        # if not self.query_obj_list:
        #     regex_results = re.findall(regex18, per_excel)
        #     if regex_results:
        #         regex_results = list(set(regex_results))  # 去重
        #     for regex_result in regex_results:
        #         device_name = regex_result[0].strip().lower()
        #         interface = regex_result[1].lower().strip().replace(" ", "").replace("e", "").replace("xg","ten")
        #         objs_switch = models.Swtich_Device.objects.filter(name__contains=device_name)
        #         if objs_switch.count() == 1:
        #             obj_switch = objs_switch[0]
        #             obj_switch.query_type = "switch"
        #             obj_switch.version = obj_switch.version
        #             obj_switch.ip = obj_switch.manege_ip
        #             obj_switch.interfaces = interface  # 将要查询的端口插入到对象中
        #             obj_switch.keywords = regex_result
        #             obj_switch.num = num
        #
        #             self.query_obj_list.append(obj_switch)

        # 匹配配置信息
        if not self.query_obj_list:
            regex_results = re.findall(regex3, self.row_excel)  # 匹配配置信息
            if regex_results:
                # print("匹配配置信息：", ret1)
                regex_results = list(set(regex_results))  # 去重
                # print("匹配到配置信息:",ret)
            for regex_result in regex_results:
                """这里进行2次正则匹配是故意的"""
                ret = re.findall(r"^(.+?)\..+?\.(\d{2})", regex_result)
                # -----------这里是ONU的部分--------------------------------
                for olt_name, onu_interface in ret:
                    objs_olt = models.Olt.objects.filter(name=olt_name.lower())
                    for obj_olt in objs_olt:
                        onu_interface = "%s/%s" % (
                            obj_olt.interface,
                            int(onu_interface)) if obj_olt.epon.tag.version.name == "ISCOM5800E-SMCB_1.44" \
                            else "onu%s:%s" % (obj_olt.interface[3:], int(onu_interface))
                        objs_onu = models.Onu.objects.filter(interface=onu_interface, olt=obj_olt)
                        for obj_onu in objs_onu:
                            obj_onu.query_type = "onu"
                            obj_onu.version = obj_olt.epon.tag.version.name
                            obj_onu.ip = obj_olt.epon.IP
                            obj_onu.interfaces = onu_interface  # 将要查询的端口插入到对象中
                            obj_onu.keywords = regex_result
                            obj_onu.num = self.num

                            self.query_obj_list.append(obj_onu)

                # -----------这里是交换机的部分--------暂时放着------------------------
                # 匹配交换机端口的配置信息还是不好做
                # else:  # 还是有部分VPN是VIP交换机接入的#
                #     ret = re.findall(r"(\d{4}[VIvi][-_][A-Za-z]).*?\..+?\.([\w /]*)", rmss_code)
                #     if ret:
                #         aaa, interface = ret[0]
                #         switch_objs = models.Swtich_Device.objects.filter(name__contains=aaa.lower())
                #         for switch_obj in switch_objs:
                #             switch_obj.interface = interface.lower().replace("e", "")
                #             switch_obj.type = "switch"
                #             switch_obj.num = num
                #             switch_obj.rmss_code = rmss_code
                #             result_info_obj.objs_onu_list.append(switch_obj)

        # 匹配业务编码
        if not self.query_obj_list:
            for regex in list_regex:  # 匹配业务编码
                regex_results = re.findall(regex, self.row_excel)
                if regex_results:
                    regex_results = list(set(regex_results))
                for regex_result in regex_results:
                    rmss_code_list = get_busenesses_brief_info_by_businesses_code(regex_result)
                    for rmss_code in rmss_code_list:
                        ret = re.findall(r"^(.+?)\..+?\.(\d{2})", rmss_code)
                        # print("爬虫爬到的配置信息：%s，匹配onu的匹配结果：%s"%(rmss_code,ret))
                        # -----------这里是ONU的部分--------------------------------
                        if ret:
                            olt_name, onu_interface = ret[0]
                            objs_olt = models.Olt.objects.filter(name=olt_name.lower())
                            for obj_olt in objs_olt:
                                onu_interface = "%s/%s" % (
                                    obj_olt.interface,
                                    int(onu_interface)) if obj_olt.epon.tag.version.name == "ISCOM5800E-SMCB_1.44" \
                                    else "onu%s:%s" % (obj_olt.interface[3:], int(onu_interface))
                                objs_onu = models.Onu.objects.filter(olt=obj_olt, interface=onu_interface)
                                for obj_onu in objs_onu:
                                    obj_onu.query_type = "onu"
                                    obj_onu.version = obj_olt.epon.tag.version.name
                                    obj_onu.ip = obj_olt.epon.IP
                                    obj_onu.interfaces = onu_interface  # 将要查询的端口插入到对象中
                                    obj_onu.keywords = regex_result
                                    obj_onu.num = self.num

                                    self.query_obj_list.append(obj_onu)

                        # #-----------这里是交换机的部分--------------------------------
                        # """待修改，因为switch对象没做"""
                        else:
                            # regex_results = re.findall(regex22, rmss_code)
                            regex_results = re.findall(r"(\d{4}[VIvi][-_][A-Za-z]).*?\..+?\.([\w /]*)", rmss_code)
                            if regex_results:
                                regex_results = list(set(regex_results))  # 去重
                            for device_name,interface in regex_results:
                                device_name = device_name.strip().lower()
                                interface = interface.lower().strip().replace(" ", "").replace("e", "").replace("xg","ten")
                                objs = models.Device.objects.filter(name__contains=device_name)
                                # print(objs)
                                if objs.count() == 1:
                                    obj = objs[0]
                                    obj.query_type = "ISP_Access"
                                    obj.version = obj.tag.version.name
                                    obj.ip = obj.IP
                                    obj.interfaces = interface  # 将要查询的端口插入到对象中
                                    obj.keywords = regex_result
                                    obj.num = self.num

                                    self.query_obj_list.append(obj)

        # 匹配分光器编号
        if not self.query_obj_list:
            for regex in list_regex2:  # 匹配分光器编号
                regex_results = re.findall(regex, self.row_excel)
                if regex_results:
                    # print("分光器编号的匹配结果：", regex_results)
                    regex_results = list(set(regex_results))
                for regex_result in regex_results:
                    objs_olt = models.Olt.objects.filter(name=regex_result.lower())
                    # 这里使用包含的话，有时会找到过多，这肯定更不好。。
                    for obj_olt in objs_olt:
                        obj_olt.query_type = "olt"
                        obj_olt.version = obj_olt.epon.tag.version.name
                        obj_olt.ip = obj_olt.epon.IP
                        obj_olt.interfaces = obj_olt.interface  # 将要查询的端口插入到对象中
                        obj_olt.keywords = regex_result
                        obj_olt.num = self.num

                        self.query_obj_list.append(obj_olt)

                if self.query_obj_list:
                    break  # 这里可能会有小bug，先忽略吧；会导致匹配一个分光器有结果后就会退出

        # return self.query_obj_list, regex_result #其实无所谓返回信息，把该保存的信息都已经保存到self中了
        # print(4567,self.query_obj_list)


class FinalResult:
    """用于保存最终结果"""

    def __init__(self):
        self.rows_list = []  # 查询对象按excel行内容进行排列
        self.queried_obj_sored_by_IP = []  # 列表内包含字典；查询对象按管理ip进行排列

        self.command_result_detail = ""  # 字符串，写入表二的查询结果
        self.previous_onu_mac_result = None  # 已经json.load，列表对象
        self.curr_onu_mac_result = None  # 与上面的对应，包含了当前olt下全部onu mac和在线onu mac
        self.json_curr_onu_mac_result = None  # json字符串，用于写入表3，self.curr_onu_mac_result.dumps处理结果
        self.curr_online_onu_mac_count_list = []
        self.curr_onu_mac_count_list = []
        self.content_offline_summary = ""  # 字符串，用于写入表4离线onu概况
        self.content_offline_detail = ""  # 字符串，用于写入表5离线onu详情

        self.wb = openpyxl.Workbook()  # 创建excel对象
        self.path = ""
        self.font1 = Font(size=9, )
        self.align = Alignment(horizontal='center')
        self.num_format = '0.00%'
        self.side = Side(style='thin', color="000000")
        self.border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)

    """下面的方法用于处理前端传递的割接内容，生成2个列表self.rows_list、self.queried_obj_sored_by_IP"""

    def __create_rows_list(self, excel_content):
        rows_excel = excel_content.split("\r\n")
        pool = ThreadPoolExecutor(INT1)  # 启动多线程查询excel表行内容
        for index, row_excel in enumerate(rows_excel, 1):
            pool.submit(self.__add_into_rows_list, index, row_excel)
        pool.shutdown(wait=True)

    def __add_into_rows_list(self, index, row_excel):
        row_obj = Row(index, row_excel)
        row_obj.get_query_objs()
        self.rows_list.append(row_obj)

    def __get_queried_obj_sored_by_IP(self):
        """将所有查询obj根据ip进行排列，返回一个包含字典的列表，保存到self.queried_obj_sored_by_IP中"""
        list1 = []
        objs_list_by_IP = []

        for row_obj in self.rows_list:
            for query_obj in row_obj.query_obj_list:
                if query_obj.ip not in list1:
                    dict1 = {"IP": query_obj.ip, "query_objs": [query_obj, ]}
                    list1.append(query_obj.ip)
                    objs_list_by_IP.append(dict1)
                else:
                    for dict1 in objs_list_by_IP:
                        if dict1["IP"] == query_obj.ip:
                            dict1["query_objs"].append(query_obj)
        self.queried_obj_sored_by_IP = objs_list_by_IP

    def processing_excel_content(self, excel_content):
        self.__create_rows_list(excel_content)
        self.__get_queried_obj_sored_by_IP()  # 将所有的查询对象根据ip进行排列，保存到对象字段中
        self.rows_list.sort(key=lambda x: x.num, reverse=False)

    """下面的方法用于登录设备进行查询，并生成表2的数据"""

    def __telnet_query_sub(self, username, password, index, dict1):
        username = bytes(username, "utf-8")
        password = bytes(password, "utf-8")
        IP = dict1["IP"]
        version = dict1["query_objs"][0].version
        query_type = dict1["query_objs"][0].query_type
        query_objs = dict1["query_objs"]
        print(index, "正在登陆%s进行查询" % IP)
        try:
            if query_type == "onu" or query_type == "olt":  # 这里处理的是EPON设备的割接
                epon_obj = network_device.Epon(IP)
                epon_obj.device_login(username, password)
                for query_obj in query_objs:
                    if query_obj.query_type == 'olt':
                        if version == "Version 5.20":
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"dis onuinfo inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content  # 保存了该对象输入的所有命令的结果
                            query_obj.onu_mac_count = re.findall(r"\w{4}-\w{4}-\w{4}", per_content)  # 保存了所有onu的mac
                            query_obj.online_onu_mac_count = re.findall(
                                r"(\w{4}-\w{4}-\w{4}).*?(Onu\d{1,2}/0/\d{1,2}:\d{1,2}).*?Up",
                                per_content)  # 保存了所有在线onu的mac
                            count = len(query_obj.onu_mac_count)
                            online_count = len(query_obj.online_onu_mac_count)
                            query_obj.result_brief = "onu共%s台，在线%s台；" % (count, online_count)
                        elif version == "Version 7.1":
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail = per_content
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"dis onu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content  # 保存了该对象输入的所有命令的结果
                            query_obj.onu_mac_count = re.findall(r"\w{4}-\w{4}-\w{4}", per_content)  # 保存了所有onu的mac
                            query_obj.online_onu_mac_count = re.findall(
                                r"(\w{4}-\w{4}-\w{4}).*?(Onu\d{1,2}/0/\d{1,2}:\d{1,2}).*?Up",
                                per_content)  # 保存了所有在线onu的mac
                            count = len(query_obj.onu_mac_count)
                            online_count = len(query_obj.online_onu_mac_count)
                            query_obj.result_brief = "onu共%s台，在线%s台；" % (count, online_count)
                        else:
                            epon_obj.tn.write(
                                b"show running-config  inter port  %s \n" % port_mapping_reverse.get(
                                    query_obj.interfaces,
                                    "0").encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail = per_content
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"show int olt %s inf   \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            epon_obj.tn.write(b"show int onu inf | inc %s/   \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            query_obj.onu_mac_count = re.findall(r"\w{4}\.\w{4}\.\w{4}", per_content)
                            query_obj.online_onu_mac_count = re.findall(
                                r"(\d{1,2}/\d{1,2}/\d{1,2}).*?(\w{4}\.\w{4}\.\w{4}) *?online", per_content)
                            count = len(query_obj.onu_mac_count)
                            online_count = len(query_obj.online_onu_mac_count)
                            query_obj.result_brief = "onu共%s台，在线%s台；" % (count, online_count)

                            list_online_onu_mac_reverse = []
                            for www in query_obj.online_onu_mac_count:  # 瑞斯的onu mac、端口的顺序与华三相反，这里进行重新排序
                                www = list(www)
                                www.reverse()
                                list_online_onu_mac_reverse.append(www)
                            query_obj.online_onu_mac_count = list_online_onu_mac_reverse
                    else:  # 这里是查询onu
                        if version == "Version 5.20":
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.olt.interface.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail += per_content
                            epon_obj.tn.write(b"dis onuinfo inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            onu_status = "在线" if "Up" in per_content else "离线"
                            epon_obj.tn.write(b"dis mac-addr int %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            mac_status = "有mac" if re.findall(r"\w{4}-\w{4}-\w{4}", per_content) else "无mac"
                            query_obj.result_brief = "onu%s，%s" % (onu_status, mac_status)
                        elif version == "Version 7.1":
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.olt.interface.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"dis cu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail += per_content
                            epon_obj.tn.write(b"dis onu inter %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            onu_status = "在线" if "Up" in per_content else "离线"
                            epon_obj.tn.write(b"dis mac-addr int %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail += per_content
                            mac_status = "有mac" if re.findall(r"\w{4}-\w{4}-\w{4}", per_content) else "无mac"
                            query_obj.result_brief = "onu%s，%s" % (onu_status, mac_status)
                        else:  # "Raisecom"
                            epon_obj.tn.write(
                                b"show running-config  inter port  %s \n" % port_mapping_reverse.get(
                                    query_obj.olt.interface,
                                    "0").encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"show running-config  onu  %s \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.desc = re.findall(r"description (\S+)", per_content)[0] if re.findall(
                                r"description (\S+)", per_content) else "无描述or匹配异常"
                            query_obj.command_result_detail = per_content
                            epon_obj.tn.write(b"show int onu %s inf   \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail = per_content
                            onu_status = "在线" if "online" in per_content else "离线"
                            epon_obj.tn.write(b"show int onu %s mac-addr l2 dy   \n" % query_obj.interfaces.encode())
                            time.sleep(0.1)
                            per_content = mydecode(my_read_very_eager(epon_obj.tn))
                            query_obj.command_result_detail = per_content
                            mac_status = "有mac" if re.findall(r"\w{4}\.\w{4}\.\w{4}.*? [123456789] .*?\d{1,3}",
                                                              per_content) else "无mac"
                            query_obj.result_brief = "onu%s，%s" % (onu_status, mac_status)
                    # print("line414,telnet后，query_obj的状态,：",query_obj.__dict__)
                if version == "ISCOM5800E-SMCB_1.44":
                    epon_obj.tn.write(b"terminal page-break enable \n")
                    print("瑞斯设备,割接查询，开启分页显示，", IP)
            else:  # -------------所有的非EPON查询都可以写在这里------
                """在非epon查询中，query_type值就是Tag.name：POP/BRAS/PE等等
                设计了用于登录设备的类的类名与Tag.name同名"""
                Device_Class = getattr(network_device, query_type)  # 根据query_type值到network_device模块中取类
                if Device_Class:
                    obj_device = Device_Class(IP)  # 创建对象时，会自动telnet连接到该设备
                    obj_device.device_login(username, password)  # 登录该设备
                    for query_obj in query_objs:
                        obj_device.cutting_query(query_obj)  # 使用query_obj对象中保存的端口信息进行查询，并将查询结果存放在query_obj对象中
                    obj_device.close()

                # if query_type == "CMTS_POP":
                #     obj_device = CMTS_POP(IP)#创建对象时，会自动telnet连接到该设备
                #     obj_device.device_login(username,password)#登录该设备
                #     for query_obj in query_objs:
                #         obj_device.cutting_query(query_obj)#使用query_obj对象中保存的端口信息进行查询，并将查询结果存放在query_obj对象中
                # elif query_type == "BRAS":
                #     obj_device = BRAS(IP)
                #     obj_device.device_login(username,password)
                #     for query_obj in query_objs:
                #         obj_device.cutting_query(query_obj)

        except:
            content = "设备登录异常，无法查询，请重试"  # 这里要重写点什么呢？？
        finally:
            try:
                epon_obj.close()
                print("割接查询结束，关闭连接，", IP)
            except:
                pass

    def __get_command_result_detail(self):
        for result_obj in self.rows_list:
            self.command_result_detail += "割接表原始内容：" + result_obj.row_excel + "\r\n"

            for query_obj in result_obj.query_obj_list:
                self.command_result_detail += "匹配结果（用于查询的依据）：" + str(query_obj.keywords) + "\r\n"
                self.command_result_detail += "匹配结果（查询结果）：%s %s %s   \r\n" % (
                    query_obj.ip, query_obj.interfaces, getattr(query_obj, "result_brief", "WRONG"))
                self.command_result_detail += getattr(query_obj, "command_result_detail",
                                                      "WRONG WRONG WRONG WRONG WRONG WRONG WRONG ")
            self.command_result_detail += "\r\n\r\n\r\n" + "*" * 150 + "\r\n\r\n\r\n\r\n"

    def telnet_query(self, username, password):
        pool = ThreadPoolExecutor(INT2)
        for num, dict1 in enumerate(self.queried_obj_sored_by_IP):  # 然后开始登陆设备进行查询，所有的详细结果都保存在最小单位query_obj中
            pool.submit(self.__telnet_query_sub, username, password, num, dict1)
        pool.shutdown(wait=True)

        self.__get_command_result_detail()

    """下面的方法用于生成表3的数据。同时前端若有传递上次的查询结果，则进行处理，生成表4,5的数据；没有则忽略"""

    def __get_previous_onu_mac_result(self, previous_online_onu_mac_result):
        if previous_online_onu_mac_result:
            # previous_online_onu_mac_result = previous_online_onu_mac_result.replace("\r\n", "")
            previous_online_onu_mac_result = previous_online_onu_mac_result.replace("\t", "").replace("\r\n", "")
            try:
                self.previous_onu_mac_result = json.loads(previous_online_onu_mac_result)
            except Exception as error1:
                print("ERROR:", error1)

    def __get_json_onu_mac_result(self):
        self.curr_onu_mac_result = []  # json序列化的就是这个对象；这个表内按查询顺序储存列表，列表内包含字典
        for row_obj in self.rows_list:
            per_online_onu_mac_result = []
            for obj in row_obj.query_obj_list:
                if obj.query_type == "olt":
                    dict3 = {"device_ip": obj.ip,
                             # "onu_mac_count":obj.onu_mac_count,
                             "onu_mac_count": getattr(obj, "onu_mac_count", "WRONG"),
                             # "list_online_onu_mac":obj.online_onu_mac_count,
                             "list_online_onu_mac": getattr(obj, "online_onu_mac_count", "WRONG"),
                             "regex_result": obj.keywords,
                             "olt_interface": obj.interfaces,
                             "version": obj.version,
                             }
                    per_online_onu_mac_result.append(dict3)
            self.curr_onu_mac_result.append(per_online_onu_mac_result)
        self.json_curr_onu_mac_result = json.dumps(self.curr_onu_mac_result)

    def __get_current_onu_online_offline_info(self):  # 返回2个列表，列表内分别包含在线mac和左右mac

        for layer1_list in self.curr_onu_mac_result:
            for layer2_dict in layer1_list:
                for i in layer2_dict.get("list_online_onu_mac", ""):
                    self.curr_online_onu_mac_count_list.append(i[0])
                for i in layer2_dict.get("onu_mac_count", ""):
                    self.curr_onu_mac_count_list.append(i)

    def __get_offline_info(self, username, password):
        username = bytes(username, "utf-8")
        password = bytes(password, "utf-8")
        for i in self.previous_onu_mac_result:
            for w in i:
                device_ip = w["device_ip"]
                olt_interface = w["olt_interface"]
                version = w["version"]
                list_online_onu_mac = w["list_online_onu_mac"]

                offline_onu_list = []
                for onu_mac_and_interface in list_online_onu_mac:
                    if onu_mac_and_interface[0] in self.curr_onu_mac_count_list and onu_mac_and_interface[0] \
                            not in self.curr_online_onu_mac_count_list:
                        # 上一次在线的onu，若不在本次在线onu名单内，且在本次查询的所有onu名单内；则判断为该onu离线
                        offline_onu_list.append(onu_mac_and_interface)
                if offline_onu_list:
                    print("登陆%s，进行onu离线时间查询......" % device_ip)
                    self.content_offline_summary += "割接匹配依据：%s \n" % i[0].get("regex_result")
                    self.content_offline_detail += "割接匹配依据：%s \n" % i[0].get("regex_result")
                    self.content_offline_summary += "管理ip：%s,PON口：%s \n" % (device_ip, olt_interface)
                    self.content_offline_detail += "管理ip：%s,PON口：%s \n" % (device_ip, olt_interface)
                    try:
                        epon_obj = network_device.Epon(device_ip)
                        epon_obj.device_login(username, password)
                        if version == "Version 5.20":
                            for offline_onu in offline_onu_list:
                                epon_obj.tn.write(b'dis onu-event inter %s \n' % offline_onu[1].encode())
                                content1 = my_read_very_eager(epon_obj.tn)
                                content1 = mydecode(content1)
                                ret = re.findall(r"\d{4}/\d{2}/\d{2} *?\d{2}:\d{2}:\d{2}", content1)
                                if ret:
                                    last_offline_time = ret[-1]
                                    self.content_offline_summary += "ONU编号：%s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += "ONU编号：%s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += content1
                        elif version == "Version 7.1":
                            for offline_onu in offline_onu_list:
                                epon_obj.tn.write(b'dis epon onu-event inter %s \n' % offline_onu[1].encode())
                                content1 = my_read_very_eager(epon_obj.tn)
                                content1 = mydecode(content1)
                                ret = re.findall(r"\d{4}/\d{2}/\d{2} *?\d{2}:\d{2}:\d{2}", content1)
                                if ret:
                                    last_offline_time = ret[-1]
                                    self.content_offline_summary += "ONU编号：%s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += "ONU编号：%s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += content1
                        else:  # Raisecom
                            for offline_onu in offline_onu_list:
                                epon_obj.tn.write(b'show alarm log int onu %s \n' % offline_onu[1].encode())
                                content1 = my_read_very_eager(epon_obj.tn)
                                content1 = mydecode(content1)
                                ret = re.findall(r"\w{2,4}-\d{2}-\d{4} *?\d{2}:\d{2}:\d{2}", content1)
                                if ret:
                                    last_offline_time = ret[-1]
                                    self.content_offline_summary += "ONU编号：onu %s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += "ONU编号：%s,ONU MAC:%s,最后离线时间：%s \n" % (
                                        offline_onu[1], offline_onu[0], last_offline_time)
                                    self.content_offline_detail += content1
                            epon_obj.tn.write(b"terminal page-break enable \n")
                            print("瑞斯设备表4,5查询，开启分页", device_ip)
                    except:
                        pass
                    finally:
                        try:
                            epon_obj.tn.close()
                            print("割接表45查询结束，关闭连接", device_ip)
                        except:
                            pass
                    self.content_offline_summary += "\r\n\r\n\r\n" + "*" * 150 + "\r\n\r\n\r\n\r\n"
                    self.content_offline_detail += "\r\n\r\n\r\n" + "*" * 150 + "\r\n\r\n\r\n\r\n"

    def processing_data_for_excel345(self, previous_online_onu_mac_result, username, password):
        self.__get_previous_onu_mac_result(previous_online_onu_mac_result)
        self.__get_json_onu_mac_result()
        if self.previous_onu_mac_result:
            self.__get_current_onu_online_offline_info()
            self.__get_offline_info(username, password)

    """下面的方法用于将数据下入excel表中"""

    def __get_file_path(self):
        path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.path = os.path.join(path, "statics", "cutting_query_result", "割接查询结果.xlsx")

    def __write_into_excel1(self):
        max_num = 0
        for row_obj in self.rows_list:  # i:[[[管理ip，端口，在线情况],[管理ip，端口，在线情况]]，匹配依据，excel表行内容]
            num = len(row_obj.query_obj_list)
            if num > max_num:
                max_num = num

        sheet = self.wb.active
        sheet.title = '割接查询汇总'

        import datetime
        sheet.cell(row=1, column=1, value="查询时间:%s" % datetime.datetime.now())
        for i in range(max_num):
            sheet.cell(row=2, column=5 * i + 1, value="割接查询依据")
            sheet.cell(row=2, column=5 * i + 2, value="端口当前描述")
            sheet.cell(row=2, column=5 * i + 3, value="管理ip")
            sheet.cell(row=2, column=5 * i + 4, value="端口")
            sheet.cell(row=2, column=5 * i + 5, value="当前在线情况")
        sheet.cell(row=2, column=max_num * 5 + 1, value="原始数据")

        for num, row_obj in enumerate(self.rows_list):  # 写入割接查询汇总表内
            for num2, query_obj in enumerate(row_obj.query_obj_list):
                try:
                    a = sheet.cell(row=num + 3, column=num2 * 5 + 1, value=str(query_obj.keywords))
                    a.border = self.border
                    a.font = self.font1
                    a = sheet.cell(row=num + 3, column=num2 * 5 + 2, value=getattr(query_obj, "desc", "无描述"))
                    a.border = self.border
                    a.font = self.font1
                    a = sheet.cell(row=num + 3, column=num2 * 5 + 3, value=query_obj.ip)
                    a.border = self.border
                    a.font = self.font1
                    if getattr(query_obj, "epon", None) and query_obj.epon.tag.version.name == "ISCOM5800E-SMCB_1.44":
                        a = sheet.cell(row=num + 3, column=num2 * 5 + 4, value="olt" + str(query_obj.interfaces))
                    elif getattr(query_obj, "olt",
                                 None) and query_obj.olt.epon.tag.version.name == "ISCOM5800E-SMCB_1.44":
                        a = sheet.cell(row=num + 3, column=num2 * 5 + 4, value="onu" + str(query_obj.interfaces))
                    else:
                        a = sheet.cell(row=num + 3, column=num2 * 5 + 4, value=query_obj.interfaces)
                    a.border = self.border
                    a.font = self.font1
                    a = sheet.cell(row=num + 3, column=num2 * 5 + 5, value=getattr(query_obj, "result_brief", "WRONG"))
                    a.border = self.border
                    a.font = self.font1
                except Exception as error1:
                    print("写入excel表1时出错，ERROR:", error1)
            a = sheet.cell(row=num + 3, column=max_num * 5 + 1, value=row_obj.row_excel)
            a.border = self.border
            a.font = self.font1

    def __write_into_excel2(self):
        sheet2 = self.wb.create_sheet('割接查询详情', index=1)
        command_result_detail = self.command_result_detail.replace("\x1b[2J", "").replace("\x1b[0;0H\r", "")
        command_result_detail_list = command_result_detail.split("\n")

        for num, i in enumerate(command_result_detail_list):  # 写入割接查询详情表内
            try:
                a = sheet2.cell(row=num + 1, column=1, value=i)
            except:
                try:
                    i = i.encode("utf-8")
                    i = i.replace(b"\x06", b"").replace(b"\x05", b"").replace(b"\x07", b"")
                    i = i.decode("utf-8")
                    # a = sheet2.cell(row=num + 1, column=1, value=i)
                    a = sheet2.cell(row=num + 1, column=1, value="该行数据编码问题，写入excel时发生错误，已被本语句替换，若影响割接查询，请登录设备进行查看")
                    a.font = self.font1
                except:
                    pass

    def __write_into_excel3(self):
        sheet3 = self.wb.create_sheet('查询结果（JSON字符串）', index=2)
        length = len(self.json_curr_onu_mac_result) // 30000 + 1
        for i in range(length):
            a = sheet3.cell(row=i + 1, column=1, value=self.json_curr_onu_mac_result[i * 30000:(1 + i) * 30000])
            a.font = self.font1

    def __write_into_excel45(self):
        if self.previous_onu_mac_result:
            sheet4 = self.wb.create_sheet('离线onu概况', index=3)
            sheet5 = self.wb.create_sheet('离线onu详情', index=4)
            content_offline_summary_list = self.content_offline_summary.split("\n")
            for num, i in enumerate(content_offline_summary_list):
                try:
                    a = sheet4.cell(row=num + 1, column=1, value=i)
                    a.font = self.font1
                except:
                    i = i.encode("utf-8")
                    i = i.replace(b"\x06", b"").replace(b"\x05", b"").replace(b"\x07", b"")
                    i = i.decode("utf-8")
                    a = sheet4.cell(row=num + 1, column=1, value=i)
                    a.font = self.font1
            content_offline_detail_list = self.content_offline_detail.split("\n")
            for num, i in enumerate(content_offline_detail_list):
                try:
                    a = sheet5.cell(row=num + 1, column=1, value=i)
                    a.font = self.font1
                except:
                    i = i.encode("utf-8")
                    i = i.replace(b"\x06", b"").replace(b"\x05", b"").replace(b"\x07", b"")
                    i = i.decode("utf-8")
                    a = sheet5.cell(row=num + 1, column=1, value=i)
                    a.font = self.font1

    def save_into_file(self):
        self.__get_file_path()
        self.__write_into_excel1()
        self.__write_into_excel2()
        self.__write_into_excel3()
        self.__write_into_excel45()
        self.wb.save(self.path)